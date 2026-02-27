import openpyxl
from openpyxl.styles import PatternFill

from excel_handling import workbook


def get_row_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def get_column_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def readData(file,sheetname,rownumber,columnnumber):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownumber,columnnumber).value

def writeData(file,sheetname,rownumber,columnnumber,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownumber, columnnumber).value=data
    workbook.save(file)

def green_fill(file,sheetname,rownumber,columnnumber):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greencolor=PatternFill(start_color="00ff00",fill_type="solid")
    sheet.cell(rownumber,columnnumber).fill=greencolor
    workbook.save(file)

def red_fill(file,sheetname,rownumber,columnnumber):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    redcolor=PatternFill(start_color="ff0000",fill_type="solid")
    sheet.cell(rownumber,columnnumber).fill=redcolor
    workbook.save(file)








