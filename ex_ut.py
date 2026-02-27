from selenium import webdriver
import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def get_column_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def get_row_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def readData(file,sheetname,rownumber,columnnumber):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownumber,columnnumber).value

def writeData(file,sheetname,rownumber,columnnumber,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownumber,columnnumber).value=data
    workbook.save()





