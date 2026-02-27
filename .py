from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from time import sleep

# === Excel Utility Functions ===

def get_row_count(file, sheetname):
    wb = load_workbook(file)
    sheet = wb[sheetname]
    return sheet.max_row

def get_column_count(file, sheetname):
    wb = load_workbook(file)
    sheet = wb[sheetname]
    return sheet.max_column

def readData(file, sheetname, row, col):
    wb = load_workbook(file)
    sheet = wb[sheetname]
    return sheet.cell(row, col).value

def writeData(file, sheetname, row, col, data):
    wb = load_workbook(file)
    sheet = wb[sheetname]
    sheet.cell(row, col).value = data
    wb.save(file)

def green_fill(file, sheetname, row, col):
    wb = load_workbook(file)
    sheet = wb[sheetname]
    green = PatternFill(start_color="00FF00", fill_type="solid")
    sheet.cell(row, col).fill = green
    wb.save(file)

def red_fill(file, sheetname, row, col):
    wb = load_workbook(file)
    sheet = wb[sheetname]
    red = PatternFill(start_color="FF0000", fill_type="solid")
    sheet.cell(row, col).fill = red
    wb.save(file)

# === Form Filling Begins ===

file = "students.xlsx"         # Path to your Excel file
sheet = "Sheet1"               # Sheet name inside Excel

driver = webdriver.Chrome()    # Launch Chrome browser
driver.maximize_window()       # Maximize window

rows = get_row_count(file, sheet)  # Total student entries

for r in range(2, rows + 1):
    driver.get("https://testautomationpractice.blogspot.com/")
    sleep(2)

    # === Read data from Excel ===
    first = readData(file, sheet, r, 2)
    last = readData(file, sheet, r, 3)
    phone = readData(file, sheet, r, 4)
    gender = readData(file, sheet, r, 5)
    day = readData(file, sheet, r, 6)
    country = readData(file, sheet, r, 7)
    colors = readData(file, sheet, r, 8)
    date = readData(file, sheet, r, 9)

    form_ok = True  # Flag to track success

    # === Fill the form ===

    # First Name
    driver.find_element(By.ID, "name").clear()
    driver.find_element(By.ID, "name").send_keys(first)

    # Last Name
    driver.find_element(By.ID, "lastname").clear()
    driver.find_element(By.ID, "lastname").send_keys(last)

    # Phone
    driver.find_element(By.ID, "phone").clear()
    driver.find_element(By.ID, "phone").send_keys(str(phone))

    # Gender Radio
    if gender and gender.lower() == "male":
        driver.find_element(By.ID, "male").click()
    elif gender and gender.lower() == "female":
        driver.find_element(By.ID, "female").click()
    else:
        form_ok = False

    # Day Checkbox
    if day:
        driver.find_element(By.XPATH, f"//input[@type='checkbox' and @value='{day}']").click()
    else:
        form_ok = False

    # Country Dropdown
    if country:
        Select(driver.find_element(By.ID, "country")).select_by_visible_text(country)
    else:
        form_ok = False

    # Colors Multi-Select
    if colors:
        color_list = [c.strip() for c in colors.split(",")]
        color_element = driver.find_element(By.NAME, "colors")
        color_select = Select(color_element)
        for color in color_list:
            color_select.select_by_visible_text(color)
    else:
        form_ok = False

    # Date Picker
    if date:
        date_field = driver.find_element(By.ID, "datepicker")
        date_field.clear()
        date_field.send_keys(date)
    else:
        form_ok = False

    # Submit Button
    driver.find_element(By.ID, "submitbtn").click()
    sleep(2)

    # === Update Status in Excel ===
    if form_ok:
        writeData(file, sheet, r, 10, "Success")
        green_fill(file, sheet, r, 10)
    else:
        writeData(file, sheet, r, 10, "Failed")
        red_fill(file, sheet, r, 10)

# === Done with all students ===
driver.quit()
