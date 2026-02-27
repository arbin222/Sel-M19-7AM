from selenium import webdriver
import openpyxl

driver=webdriver.Chrome()
file= r"C:\Users\Levono\OneDrive\Desktop\Book1.xlsx"
workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet2"]
# rows=sheet.max_row
# cols=sheet.max_column


# same data for 5 rows and 5 columns
# for row in range(1,6):
#     for col in range(1,6):
#         sheet.cell(row,col).value="Selenium"
# workbook.save(file)
#
#
# sheet=workbook["Sheet1"]
# rows=sheet.max_row
# cols=sheet.max_column
# #
# # Different data for rows and columns
#
# sheet.cell(1,1).value="sl.no"
# sheet.cell(1,2).value="name"
# sheet.cell(1,3).value="phone_no."
# sheet.cell(1,4).value="gender"
# sheet.cell(1,5).value="course"
#
# sheet.cell(2,1).value="1"
# sheet.cell(2,2).value="arshi"
# sheet.cell(2,3).value="74382198412"
# sheet.cell(2,4).value="F"
# sheet.cell(2,5).value="Python Full Stack"
#
# sheet.cell(3,1).value="2"
# sheet.cell(3,2).value="arbi"
# sheet.cell(3,3).value="2934792367"
# sheet.cell(3,4).value="F"
# sheet.cell(3,5).value="Python Full Stack"
#
# sheet.cell(4,1).value="3"
# sheet.cell(4,2).value="sadu"
# sheet.cell(4,3).value="7251848962"
# sheet.cell(4,4).value="M"
# sheet.cell(4,5).value="Java Full Stack"
#
# sheet.cell(5,1).value="4"
# sheet.cell(5,2).value="sada"
# sheet.cell(5,3).value="5872198412"
# sheet.cell(5,4).value="F"
# sheet.cell(5,5).value="Python Full Stack"
#
# workbook.save(file)
#
# print(sheet.max_row,sheet.max_column)


#read single data
# print(sheet.cell(3,4).value)


# to get all the data in the sheet
# for row in range(1,rows+1):
#     for col in range(1,cols+1):
#         print(sheet.cell(row,col).value, end=' ')
from time import sleep

driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india/fixed-deposit-calculator-SBI-BSB001.html")
driver.maximize_window()
sleep(2)

from excel_utility import *
from selenium.webdriver.support.select import Select

rows=get_row_count(file,'Sheet1')
cols=get_column_count(file,'Sheet1')

for row in range(2,rows+1):
    p=readData(file,'Sheet1',row,2)
    roi=readData(file,'Sheet1',row,3)
    pr1=readData(file,'Sheet1',row,4)

driver.quit()


